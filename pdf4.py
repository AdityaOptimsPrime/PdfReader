import streamlit as st
import PyPDF2 as pPyPDFReaderdf
import openpyxl as xl
import re
import pandas as pd
import pdfplumber as pdfp
import camelot
import os



def extract_fields_from_pdf(pdf_path):
    tables = camelot.read_pdf(filepath=pdf_path, flavor='stream', pages='all')
    for i,table in enumerate(tables):
        df = table.df
        st.write(df)
        if df.shape[1]<3:
            continue
        if df.shape[1]>3:
            df=df[[df.columns[2],df.columns[3]]]
            df.columns=['Ord    Ship    Cancel','Part Number']
            df=df[
                (df['Ord    Ship    Cancel']!='Ord    Ship    Cancel') &
                (df['Ord    Ship    Cancel']!='')&
                (df['Part Number']!='')&
                (df['Part Number']!='Part Number')
            ]
            st.write(df)

def commonholstein(pdf_path,pageCnt):
    tables = camelot.read_pdf(filepath=pdf_path, flavor='stream', pages='all')
    for i,table in enumerate(tables):
        df = table.df
        if pageCnt>1:
            if df.shape[1]==2:
                invoice=df.iloc[1,1]
                date=df.iloc[4,1]
                po=df.iloc[6,1]
        else:
            if df.shape[1]==0:
                invoice=df.iloc[2,5]
                date=df.iloc[6,5]
                po=df.iloc[9,5]
    return invoice,date,po


def holsteinPdf1(pdf_path):
    tables = camelot.read_pdf(filepath=pdf_path, flavor='stream', pages='all')
    allDicts=[]
    for i,table in enumerate(tables):
        df = table.df
        if df.shape[1]==5:
            continue
        if df.shape[1]>3:
            df=df[[df.columns[2],df.columns[4]]]
            df.columns=['Ord    Ship    Cancel','Part Number']
            df=df[
                (df['Ord    Ship    Cancel']!='Ord    Ship    Cancel') &
                (df['Ord    Ship    Cancel']!='')&
                (df['Part Number']!='')&
                (df['Part Number']!='Part Number')
            ]
            data_dict = dict(zip(df['Part Number'], df['Ord    Ship    Cancel']))
            allDicts.append(data_dict)
    return allDicts


def holsteinPdf2(filePath1):
    allDicts=[]
    tables = camelot.read_pdf(filepath=filePath1, flavor='stream', pages='all')
    for i,table in enumerate(tables):
        df = table.df
        if df.shape[1]==5:
            continue
        if df.shape[1]>3:
            df=df[[df.columns[2],df.columns[3]]]
            df.columns=['Ord    Ship    Cancel','Part Number']
            df=df[
                (df['Ord    Ship    Cancel']!='Ord    Ship    Cancel') &
                (df['Ord    Ship    Cancel']!='')&
                (df['Part Number']!='')&
                (df['Part Number']!='Part Number')
            ]
            data_dict = dict(zip(df['Part Number'], df['Ord    Ship    Cancel']))
            allDicts.append(data_dict)
    return allDicts

    #     if i==0:
    #         invoiceNo=df.iloc[2,1]
    #         invoiceDate=df.iloc[4,1]
    #         continue
    #     if df.shape[1] >= 6: 
    #         df.columns = [f"col_{i}" for i in range(df.shape[1])]
    #         filtered_df = df.loc[4:, ['col_2', 'col_3']] 
    #         filtered_df.columns = ['Shipped', 'Part Number']
    #         filtered_df = filtered_df[
    #             (filtered_df['Shipped'] != '') & (filtered_df['Part Number'] != '') 
    #             & (~filtered_df['Part Number'].str.contains("Subtotal|TOTAL", na=False)) 
    #         ]
    #         data_dict = dict(zip(filtered_df['Part Number'], filtered_df['Shipped']))
    #         allDicts.append(data_dict)
    
    # with open(filePath1, 'rb') as file:
    #     pdfReader = pPyPDFReaderdf.PdfReader(file)
    #     text = ""
    #     for page_num in range(len(pdfReader.pages)):
    #         page = pdfReader.pages[page_num]
    #         text += page.extract_text()
    #     pattern1 = r"FREIGHT PPD & ADDED\s*([\d.,]+)"
    #     pattern2 = r"FREIGHT  PPD\s*([\d.,]+)"
    #     PONumber=re.findall(pattern2,text)
    #     if not PONumber:
    #         PONumber=re.findall(pattern1,text)
    # return invoiceNo,invoiceDate,PONumber[0],allDicts
        


def addToExcel(*args):
    workbook = xl.load_workbook("Data.xlsx")
    sheet = workbook.active
    sheet.append(args)
    workbook.save("Data.xlsx")


def download(filename):
    df = pd.read_excel("Data.xlsx")
    csv = df.to_csv(index=False)
    st.download_button(
        label="Download Extracted Data as CSV",
        data=csv,
        file_name=filename,
        mime="text/csv",
    )


def deleteData():
    workbook = xl.load_workbook("Data.xlsx")
    sheet = workbook.active  # Select the active sheet
    for row in sheet.iter_rows(min_row=2):  # Skip the header row (row 1)
        for cell in row:
            cell.value = None  # Clear the cell's value
    workbook.save("Data.xlsx")  # Save the updated Excel file

def get_page_count(pdf_file):
    # Load the PDF file
    reader = pPyPDFReaderdf.PdfReader(pdf_file)
    # Get the total number of pages
    page_count = len(reader.pages)
    return page_count

# Streamlit UI
uploadedFiles = st.file_uploader("Upload your PDFs...", type=["pdf"], accept_multiple_files=True)

if uploadedFiles:
    st.write("PDF(s) Uploaded Successfully")

holstein1 = st.button("Click to add data of HOLSTEIN1 vendor")
holstein2 = st.button("Click to add data of HOLSTEIN2 vendor")

if holstein1:
    deleteData()
    pdfName=[]
    if uploadedFiles:
        for uploadedFile in uploadedFiles:
            with open("temp.pdf", "wb") as f:
                f.write(uploadedFile.read())
            count=get_page_count("temp.pdf")
            try:
                invoiceNo,invoiceDate,PONumber=commonholstein("temp.pdf",count)
                st.write(invoiceNo,invoiceDate,PONumber)
                itemMap=holsteinPdf1("temp.pdf")

                for itemDict in itemMap:
                    for key, value in itemDict.items():
                        addToExcel(invoiceNo, invoiceDate, PONumber, key, value)
            except:
                pdfName.append(uploadedFile.name)
    download("holstein.csv")
    st.write(pdfName)

if holstein2:
    deleteData()
    pdfName=[]
    if uploadedFiles:
        for uploadedFile in uploadedFiles:
            with open("temp.pdf", "wb") as f:
                f.write(uploadedFile.read())
            count=get_page_count("temp.pdf")
            try:
                invoiceNo,invoiceDate,PONumber=commonholstein("temp.pdf",count)
                
                itemMap=holsteinPdf2("temp.pdf")
                st.write(itemMap)

                for itemDict in itemMap:
                    for key, value in itemDict.items():
                        addToExcel(invoiceNo, invoiceDate, PONumber, key, value)
            except:
                pdfName.append(uploadedFile.name)
    download("holstein.csv")
    st.write(pdfName)

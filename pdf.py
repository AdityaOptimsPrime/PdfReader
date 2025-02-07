import streamlit as st
import PyPDF2 as pPyPDFReaderdf
import openpyxl as xl
import re
import pdfplumber as pdfp
import pandas as pd
import camelot
import os


if not os.path.exists("Data.xlsx"):
    raise FileNotFoundError("The file 'Data.xlsx' does not exist.")

url="https://docs.google.com/spreadsheets/d/1xZvNRfY2fq6ixt8qDhymX1drmtQkJAXV1-7kHlpydJY/edit?usp=sharing"


    # for table in tables:
    #     df = table.df
        
    #     # Identify the correct columns based on their headers
    #     # Replace 'Product and Description' and 'Ship Qty' with their exact column names
    #     # Remove blank rows if necessary and reset the index
    #     st.write(df)
#    with pdfp.open(filePath) as file:
#         tables = []
#         for i in file.pages:
#             extracted_tables = i.extract_tables()
#             # table=i.extract_tables()[0][1][0]
#             tables.extend(extracted_tables)
#             st.write(tables)
#         if tables:
#             last_table = tables[3]  # Get the last table
#             df = pd.DataFrame(last_table)
#             st.write(df)
    # with pdfp.open(filePath) as pdf1:
    #     shipPartList = []
        
    #     for page in pdf1.pages:
    #         text = page.extract_text()
    #         st.write(text)
    #         lines = text.split('\n')
    #         shipPartList.extend(lines)
    #         st.write(shipPartList)
    #     df=pd.DataFrame(shipPartList)
    #     st.write(df)
            
    

        # if tables:
        #     for table in tables:
        #         # last_table = tables[-1]  # Get the last table
        #         df = pd.DataFrame(tables)  # Convert it into a DataFrame
        #         # df = df.iloc[:, [1, 8]]
        #         st.write(df)
        #         # df = df.dropna()
            
        #     codeList = []
        #     product_codes = df.iloc[1:, 0].tolist()  # Get the first column as a list
        #     quantities = df.iloc[1:, 1].tolist()  # Get the second column as a list

        #     # Iterate over product_codes and split each product code
        #     for code in product_codes:
        #         print(type(code))
        #         split_code = code.split(" ")
        #         codeList.append(split_code[0])  # Assuming you only want the first part (e.g., '6PK1145')

        #     # Ensure that all elements in codeList are strings, not lists
        #     if any(isinstance(item, list) for item in codeList):
        #         st.write("Error: codeList contains lists instead of strings")
        #         return

        #     # Creating a dictionary from the product codes and quantities
        #     product_quantity_dict = dict(zip(codeList, quantities))

        #     # Format the dictionary to a string like '6PK1145 = 2, 6PK2095 = 6'
        #     formatted_output = ', '.join([f"{key} = {value}" for key, value in product_quantity_dict.items()])

        #     # Display the output
        #     st.write("Extracted Data: ", formatted_output)
        # else:
        #     st.write("No tables found in the PDF")
    # for i in dataframes:
    #     st.write(i)

def aptPdf(filePath):
    with pdfp.open(filePath) as file:
        table1=[]
        for i in file.pages:
            extracted_tables = i.extract_tables()
            tables=extracted_tables[0][1][1]
            data1=tables.split(":")[1]
            break

        for i in file.pages:
            extracted_tables=i.extract_tables()
            data2=(extracted_tables[1][4][0])
            data3=(extracted_tables[1][4][1])
            break

        for i in file.pages:
            extracted_tables=i.extract_tables()
            table1.append(extracted_tables[2])
            break
        
        
        itemName=[]
        itemQuantity=[]
        for row in table1:
            for cell in row:
                itemName.append(cell[0])
                itemQuantity.append(cell[2])
        itemName=itemName[1:]
        itemQuantity=itemQuantity[1:]
        itemMap = dict(zip(itemName, itemQuantity))

        return data1,data2,data3,itemMap
    

def bandoPdf(filePath):

    with pdfp.open(filePath) as file:
        tables = []
        allDicts = []
        for i in file.pages:
            extracted_tables = i.extract_tables()
            tables.extend(extracted_tables)

        if tables:
            last_table = tables[-1]  # Get the last table
            df = pd.DataFrame(last_table)  # Convert it into a DataFrame
            df1=df.iloc[1,2]
            df = df.iloc[:, [1, 8]]
            df = df.dropna()
            
            codeList = []
            product_codes = df.iloc[1:, 0].tolist()  # Get the first column as a list
            quantities = df.iloc[1:, 1].tolist()  # Get the second column as a list

            # Iterate over product_codes and split each product code
            for code in product_codes:
                print(type(code))
                split_code = code.split(" ")
                codeList.append(split_code[0])  # Assuming you only want the first part (e.g., '6PK1145')

            # Ensure that all elements in codeList are strings, not lists
            if any(isinstance(item, list) for item in codeList):
                st.write("Error: codeList contains lists instead of strings")
                return
            product_quantity_dict = dict(zip(codeList, quantities))
            allDicts.append(product_quantity_dict)
        else:
            st.write("No tables found in the PDF")
    return allDicts

def bandoPdf1(filePath1):
    tables = camelot.read_pdf(filepath=filePath1, flavor='stream', pages='all')
    allDicts = []
    for table in tables:
        df = table.df
        # Identify the correct columns based on their headers
        # Replace 'Product and Description' and 'Ship Qty' with their exact column names
        # Remove blank rows if necessary and reset the index
        if df.shape[1]>=4:
            df = df[[df.columns[1], df.columns[4]]]  # Adjust indices for correct columns
            df.columns = ['Product and Description', 'Ship Qty']  # Rename columns for clarity
            df = df[
                (df['Product and Description'] != 'Product and Description') &
                (df['Ship Qty'] != 'Ship Qty') &
                (df['Product and Description'] != '') &
                (df['Ship Qty'] != '')
            ]
            st.write(df)
            # Display the filtered data (or further process it)
            data_dict = dict(zip(df['Product and Description'], df['Ship Qty']))
            allDicts.append(data_dict)
    return allDicts

def comonBandoPdf(filePath1,pageNumber):
    tables = camelot.read_pdf(filepath=filePath1)
    df = tables[3].df
    df = df[[df.columns[1]]]
    PoNumber= df.iloc[1, 0]
    with open(filePath1, 'rb') as file:
        pdfReader = pPyPDFReaderdf.PdfReader(file)  # Properly count the number of pages
        text = ""
        for page_num in range(pageNumber):
            page = pdfReader.pages[page_num]
            text += page.extract_text()
    # Patterns for extracting invoice data
    invoiceDatePattern = r"Due Date\s*(\d{1,2}/\d{1,2}/\d{2})"
    invoiceNoPattern = r"Invoice\s*(\d+-\d+)"
    
    # Search for patterns in the extracted text
    invoiceDate = re.search(invoiceDatePattern, text)
    invoiceNo = re.search(invoiceNoPattern, text)
    
    if invoiceDate and invoiceNo:
        invoiceDate = invoiceDate.group(1)
        invoiceNo = invoiceNo.group(1)
        return PoNumber, invoiceDate, invoiceNo
    else:
        return None, None, None

def bestBuyPdf(filePath):
    with pdfp.open(filePath) as pdf1:
        shipPartList = []
        
        for page in pdf1.pages:
            text = page.extract_text()
            lines = text.split('\n')
            
            for line in lines:
                # Updated regex to capture Ship Qty (second number) instead of Order Qty (first number)
                match = re.match(r"^\s*\d+\s+(\d+)\s+[A-Z]+\s+([A-Z0-9]+)", line)
                if match:
                    ship_qty = match.group(1)  # Changed to capture second number (Ship Qty)
                    part_number = match.group(2)
                    shipPartList.append((ship_qty, part_number))
        
    with open(filePath,'rb') as file:
        pdfReader = pPyPDFReaderdf.PdfReader(file)  # Make sure this is the correct import
        pageNumber = len(pdfReader.pages)
        text = ""
        for page_num in range(pageNumber):
            page = pdfReader.pages[page_num]
            text += page.extract_text()

    poNoPattern = r"P.O.No:\s*(\d+)"
    invoiceDatePattern = r" Invoice Date:\s*((?:JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)/\d{1,2}/\d{4})"
    invoiceNoPattern = r"Invoice No:\s*(\d+)"
   
    poNo = re.search(poNoPattern, text)
    invoiceDate = re.search(invoiceDatePattern, text)
    invoiceNo = re.search(invoiceNoPattern, text)
    st.write("poNo: ",poNo)
    st.write("invoiceDate: ",invoiceDate)
    st.write("invoiceNo: ",invoiceNo)
    st.write("shipPartList: ",shipPartList)
    if poNo or invoiceDate or invoiceNo:
        if poNo:
            poNo = poNo.group(1)
        if invoiceDate:
            invoiceDate = invoiceDate.group(1)
        if invoiceNo:
            invoiceNo = invoiceNo.group(1)  
        return poNo, invoiceDate, invoiceNo, shipPartList
    else:
        return None, None, None, []


def densoPdf(filePath):
    item_shipped_dict = {}
    with pdfp.open(filePath) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            # Extract raw text from the page
            text = page.extract_text()
            if text:
                for line in text.split("\n"):
                    # Use regex to extract SHIPPED and ITEM NO. data
                    match = re.match(r"^\s*(\d+)\s+\d+\s+(\d{3,4}[-\dA-Za-z]*)\s+([\w\s]+)\s+U\$\s+([\d\.]+)", line)
                    if match:
                        shipped = match.group(1)
                        item_no = match.group(2)
                        item_shipped_dict[item_no] = int(shipped)
            else:
                st.write(f"Page {page_number} contains no extractable text.")

    with pdfp.open(filePath) as pdf:
        for page in pdf.pages:
            table = page.extract_table()
            if table:
                df = pd.DataFrame(table[1:], columns=table[0])
                # Filter required columns
                invoiceDate=df.iloc[:,1].tolist()
                customerPoNo=df.iloc[:,4].tolist()
            break
    
    with open(filePath,'rb') as file:
        pdfReader=pPyPDFReaderdf.PdfReader(file)
        pageNumber= len(pdfReader.pages)
        text=""
        for page_num in range(pageNumber):
            page=pdfReader.pages[page_num]
            text+=page.extract_text()
            # st.write(text)
    invoicePattern=r"Invoice\s*(\d+)"
    invoice=re.search(invoicePattern,text)
    if invoice:
        invoice=invoice.group(1)
    
    return invoice,invoiceDate[0],customerPoNo[0],item_shipped_dict

def fcsPdf(filePath1):
    tables = camelot.read_pdf(filepath=filePath1, flavor='stream', pages='all')
    allDicts = []
    Po = None
    for i, table in enumerate(tables):
        df = table.df
        if i == 0:
            continue

        if df.shape[1] >= 3:
            df = df[[df.columns[1], df.columns[3]]]
            df.columns = ['Shipped', 'Item']
            df = df[
                (df['Shipped'] != 'Shipped') &  
                (df['Item'] != 'Item') &       
                (df['Shipped'] != '') &        
                (df['Item'] != '') &           
                (~df['Item'].str.contains("Unnamed", na=False))
            ]
            if i == 1:
                df = df.iloc[1:]
            data_dict = dict(zip(df['Item'], df['Shipped']))
            allDicts.append(data_dict)
        if i == 1:
            Po = df.iloc[1, 1] 
    with open(filePath1, 'rb') as file:
        pdfReader = pPyPDFReaderdf.PdfReader(file)
        text = ""
        for page_num in range(len(pdfReader.pages)):
            page = pdfReader.pages[page_num]
            text += page.extract_text()
            break  
    invoice_pattern = r"#INV\d+"  
    date_pattern = r"\d{1,2}/\d{1,2}/\d{4}"
    invoiceMatches = re.findall(invoice_pattern, text)
    dateMatches = re.findall(date_pattern, text)

    invoice = invoiceMatches[0] if invoiceMatches else "N/A"
    date = dateMatches[0] if dateMatches else "N/A"
    return invoice, date, Po, allDicts


def gmbPdf(filePath1):
    tables = camelot.read_pdf(filepath=filePath1, flavor='stream', pages='all')
    allDicts = []
    for i, table in enumerate(tables):
        df = table.df
        if i==0:
            invoiceNo=df.iloc[2,1]
            invoiceDate=df.iloc[4,1]
            continue
        if df.shape[1] >= 6: 
            df.columns = [f"col_{i}" for i in range(df.shape[1])]
            filtered_df = df.loc[4:, ['col_2', 'col_3']] 
            filtered_df.columns = ['Shipped', 'Part Number']
            filtered_df = filtered_df[
                (filtered_df['Shipped'] != '') & (filtered_df['Part Number'] != '') 
                & (~filtered_df['Part Number'].str.contains("Subtotal|TOTAL", na=False)) 
            ]
            data_dict = dict(zip(filtered_df['Part Number'], filtered_df['Shipped']))
            allDicts.append(data_dict)
    
    with open(filePath1, 'rb') as file:
        pdfReader = pPyPDFReaderdf.PdfReader(file)
        text = ""
        for page_num in range(len(pdfReader.pages)):
            page = pdfReader.pages[page_num]
            text += page.extract_text()
        pattern1 = r"FREIGHT PPD & ADDED\s*([\d.,]+)"
        pattern2 = r"FREIGHT  PPD\s*([\d.,]+)"
        PONumber=re.findall(pattern2,text)
        if not PONumber:
            PONumber=re.findall(pattern1,text)
    return invoiceNo,invoiceDate,PONumber[0],allDicts


def g2sPdf(filePath1):
    customer_po = None
    items_dict = {}
    with pdfp.open(filePath1) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not customer_po:
                po_match = re.search(r"CONTACT\s+([A-Z\-]+)", text)
                if po_match:
                    customer_po = po_match.group(1).strip()

    tables = camelot.read_pdf(filepath=filePath1, flavor='stream', pages='all')
    for i,table in enumerate(tables):
        df = table.df
        if i==0:
            invoiceNo=df.iloc[1,2]
            date=df.iloc[2,2]
    return invoiceNo,date,customer_po

def g2sPdf2(file_obj):
    item_dict = {}
    
    with pdfp.open(file_obj) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            lines = text.split('\n')
            
            for line in lines:
                item_match = re.match(r'^\s*(\d+)\s+([A-Z]+-[A-Z0-9]+:?.*?)\s+(\d+)\s+EA', line)
                if item_match:
                    item_num = item_match.group(1)
                    item_code = item_match.group(2).strip()
                    quantity = item_match.group(3)
                    
                    # Add to dictionary with item_code as key and quantity as value
                    item_dict[item_code] = quantity

    return item_dict

def extractText(filePath):
    with open(filePath,'rb') as file:
        pdfReader=pPyPDFReaderdf.PdfReader(file)
        pageNumber= len(pdfReader.pages)
        text=""
        for page_num in range(pageNumber):
            page=pdfReader.pages[page_num]
            text+=page.extract_text()
            st.write(text)
    # poNoPattern = r"P.O.No:\s*(\d{9})"
    # invoiceDatePattern = r"Due Date\s*(\d{1,2}/\d{1,2}/\d{2})"
    # invoiceNoPattern = r"Invoice\s*(\d+-\d+)"
   
    # # poNo = re.search(poNoPattern, text)
    # invoiceDate = re.search(invoiceDatePattern, text)
    # invoiceNo = re.search(invoiceNoPattern, text)
    # # st.write(poNo)
    # st.write(invoiceDate)
    # st.write(invoiceNo)
    # if invoiceDate and invoiceNo:
    #     # poNo = poNo.group(1)
    #     invoiceDate = invoiceDate.group(1)
    #     invoiceNo=invoiceNo.group(1)
       
    #     return invoiceDate,invoiceNo
    # else:
    #     return  None,None
    

def display_pdf_content(filePath):

     with open(filePath, 'rb') as file:
        pdf_reader = pPyPDFReaderdf(file)
        num_pages = len(pdf_reader.pages)
        text = []
        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            page_text = page.extract_text()
            text.append(page_text)
        return text



def addToExcel (*args):
    workbook = xl.load_workbook("Data.xlsx")
    sheet = workbook.active
    sheet.append(args)
    workbook.save("Data.xlsx")

def download(filename):
    df=pd.read_excel("Data.xlsx")
    csv = df.to_csv(index=False)
    st.download_button(
            label="Download Extracted Data as CSV",
            data=csv,
            file_name=filename,
            mime="text/csv",
        )

def deleteData():
        workbook = xl.load_workbook("Data.xlsx")
        sheet = workbook.active  # Select the active sheet
        for row in sheet.iter_rows(min_row=2):  # Skip the header row (row 1)
            for cell in row:
                cell.value = None  # Clear the cell's value
        workbook.save("Data.xlsx")  # Save the updated Excel file 

def get_page_count(pdf_file):
    # Load the PDF file
    reader = pPyPDFReaderdf.PdfReader(pdf_file)
    # Get the total number of pages
    page_count = len(reader.pages)
    return page_count


uploadedFile = st.file_uploader("Upload your PDFs...", type=["pdf"], accept_multiple_files=True)


if uploadedFile is not None:
    st.write("PDF Uploaded Successfully")

apt = st.button("Click to add data of APT vendor")
bando = st.button("Click to add data of bando vendor")
bestBuy = st.button("Click to add data of Best Buy Distributor-new vendor")
denso = st.button("Click to add data of DENSO vendor")
fcs = st.button("Click to add data of FCS Automotive vendor")
gmb = st.button("Click to add data of GMB vendor")
g2s = st.button("Click to add data of G2S vendor")




if bando:
    deleteData()
    pdfName=[]
    if uploadedFile is not None:
        for uploadedFile1 in uploadedFile:
            with open("temp.pdf","wb") as f:
                f.write(uploadedFile1.read())
            try:
                count=get_page_count("temp.pdf")
                PoNumber,invoiceDate,invoiceNo=comonBandoPdf("temp.pdf",count)
                if count<2:
                    product_quantity_dict=bandoPdf("temp.pdf")
                else:
                    product_quantity_dict=bandoPdf1("temp.pdf")
                for i, product_quantity_dict in enumerate(product_quantity_dict, 1):
                    for key, value in product_quantity_dict.items():
                        addToExcel(PoNumber,invoiceDate,invoiceNo,key,value)
            except:
                pdfName.append(uploadedFile1.name)
    download("BANDO.csv")
    st.write(pdfName)

if apt:
    deleteData()
    pdfName=[]
    if uploadedFile is not None:
            for uploadedFile1 in uploadedFile:
                with open("temp.pdf","wb") as f:
                    f.write(uploadedFile1.read())
                try:
                    data1,data2,data3,itemMap=aptPdf("temp.pdf")
                    for key, value in itemMap.items():
                        addToExcel(data1,data2,data3,key, value)
                except:
                    pdfName.append(uploadedFile1.name)
    download("APT.csv")
    st.write(pdfName)



# if bestBuy:
#     deleteData()
#     pdfName=[]
#     if uploadedFile is not None:
#             for uploadedFile1 in uploadedFile:
#                 with open("temp.pdf","wb") as f:
#                     f.write(uploadedFile1.read())
#                 try:
#                     poNo, invoiceDate,invoiceNo,shipPartList=bestBuyPdf("temp.pdf")
#                     for ship_qty, part_number in shipPartList:
#                         addToExcel(poNo, invoiceDate,invoiceNo, ship_qty, part_number)
#                 except:
#                     pdfName.append(uploadedFile1.name)
#     download("BestBuy.csv")
#     st.write(pdfName)


if denso:
    deleteData()
    pdfName=[]
    if uploadedFile is not None:
        for uploadedFile1 in uploadedFile:
            with open("temp.pdf","wb") as f:
                f.write(uploadedFile1.read())
            try:
                data1,data2,data3,itemMap=densoPdf("temp.pdf")
                for key, value in itemMap.items():
                    addToExcel(data1,data2,data3,key, value)
            except:
                pdfName.append(uploadedFile1.name)
    download("Denso.csv")
    st.write(pdfName)

if fcs:
    deleteData()
    pdfName=[]
    if uploadedFile is not None:
        for uploadedFile1 in uploadedFile:
            with open("temp.pdf","wb") as f:
                f.write(uploadedFile1.read())
            try:
                invoice, date, po, itemMap = fcsPdf("temp.pdf")
                for itemDict in itemMap:
                    for key, value in itemDict.items():
                        addToExcel(invoice, date, po, key, value)
            except:
                pdfName.append(uploadedFile1.name)
    download("FCS.csv")
    st.write(pdfName)


if gmb:
    deleteData()
    pdfName=[]
    if uploadedFile is not None:
        for uploadedFile1 in uploadedFile:
            with open("temp.pdf", "wb") as f:
                f.write(uploadedFile1.read())
            try:
                invoiceNo,invoiceDate,PONumber,itemMap=gmbPdf("temp.pdf")
                
                for itemDict in itemMap:
                    for key, value in itemDict.items():
                        addToExcel(invoiceNo,invoiceDate,PONumber, key, value)
            except:
                pdfName.append(uploadedFile.name)
    st.write(pdfName)
    download("GMB.csv")

if g2s:
    deleteData()
    pdfName=[]
    if uploadedFile is not None:
        for uploadedFile1 in uploadedFile:
            with open("temp.pdf", "wb") as f:
                f.write(uploadedFile1.read())
            try:
                invoiceNo,invoiceDate,PONumber=g2sPdf("temp.pdf")
                itemMap=g2sPdf2("temp.pdf")
                st.write("invoiceNo: ",invoiceNo)
                st.write("invoiceDate: ",invoiceDate)
                st.write("PONumber: ",PONumber)
                st.write("itemMap: ",itemMap)
                for key, value in itemMap.items():
                    st.write("Key: ",key)
                    st.write("Value: ",value)
                    addToExcel(invoiceNo,invoiceDate,PONumber, key, value)
            except:
                pdfName.append(uploadedFile1.name)
    st.write(pdfName)
    download("g2s.csv")

if bestBuy:
    deleteData()
    pdfName = []
    if uploadedFile is not None:
        for uploadedFile1 in uploadedFile:
            with open("temp.pdf", "wb") as f:
                f.write(uploadedFile1.read())
            try:
                invoice_no, invoice_date, po_number, items = bestBuyPdf("temp.pdf")
                
                # Process each item
                for ship_qty, part_number in items:  # Changed this line
                    addToExcel(invoice_no, invoice_date, po_number, part_number, ship_qty)
                    
                    # For debugging
                    # st.write(f"Added: Invoice: {invoice_no}, Date: {invoice_date}, PO: {po_number}, Part: {part_number}, Qty: {ship_qty}")
            except Exception as e:
                pdfName.append(uploadedFile1.name)
                st.error(f"Error processing {uploadedFile1.name}: {str(e)}")
    
    if pdfName:
        st.write("Files with errors:", pdfName)
    download("BestBuy.csv")  # Changed filename to match vendor